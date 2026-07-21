from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def gerar_excel(dados_orcamento, output_path, bdi_percentual=0):
    """Gera o Excel do orcamento.

    bdi_percentual: percentual de BDI (Beneficios e Despesas Indiretas)
    a aplicar sobre o CUSTO DIRETO total (material + mao de obra),
    cobrindo administracao, lucro, impostos e imprevistos. Se 0 (ou nao
    informado), o relatorio mostra apenas o custo direto, sem bloco de
    BDI -- assim quem nao quiser usar BDI (ex.: orcamento so pra uso
    interno) continua funcionando igual antes.
    """
    # Garante extensao .xlsx (nao mais .csv)
    xlsx_path = output_path.replace(".csv", ".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento"

    fonte_padrao = "Arial"
    azul_cabecalho = "1F4E78"
    azul_secao = "4472A8"
    verde_venda = "2E7D32"
    cinza_claro = "F2F2F2"

    # --- Cabecalho do relatorio ---
    ws.merge_cells("A1:D1")
    ws["A1"] = "OrçaObra AI — Orçamento Estimado de Obra"
    ws["A1"].font = Font(name=fonte_padrao, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=azul_cabecalho)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # --- Cabecalho da tabela ---
    linha_cabecalho = 3
    colunas = ["Item", "Quantidade", "Preço Unitário (R$)", "Total (R$)"]
    for col_idx, titulo in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_cabecalho, column=col_idx, value=titulo)
        celula.font = Font(name=fonte_padrao, size=11, bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=azul_cabecalho)
        celula.alignment = Alignment(horizontal="center", vertical="center")

    borda_fina = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # --- Agrupa os itens por Tipo (Material / Mão de Obra), preservando
    # a ordem de chegada dentro de cada grupo. Itens sem "Tipo" (ex.:
    # relatorios antigos, ou uso direto da funcao sem essa chave) caem
    # num grupo generico "Itens" pra nao quebrar nada.
    grupos = {}
    ordem_grupos = []
    for item in dados_orcamento:
        tipo = item.get("Tipo", "Itens")
        if tipo not in grupos:
            grupos[tipo] = []
            ordem_grupos.append(tipo)
        grupos[tipo].append(item)

    linha_atual = linha_cabecalho + 1
    linhas_subtotal = []

    for tipo in ordem_grupos:
        itens_grupo = grupos[tipo]

        # --- Cabecalho da secao (ex.: "MATERIAIS", "MÃO DE OBRA") ---
        ws.merge_cells(f"A{linha_atual}:D{linha_atual}")
        celula_secao = ws.cell(row=linha_atual, column=1, value=tipo.upper())
        celula_secao.font = Font(name=fonte_padrao, size=11, bold=True, color="FFFFFF")
        celula_secao.fill = PatternFill("solid", fgColor=azul_secao)
        celula_secao.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        linha_atual += 1

        primeira_linha_grupo = linha_atual
        for i, item in enumerate(itens_grupo):
            zebra = cinza_claro if i % 2 == 1 else "FFFFFF"

            ws.cell(row=linha_atual, column=1, value=item["Material"])
            ws.cell(row=linha_atual, column=2, value=item["Quantidade"])
            ws.cell(row=linha_atual, column=3, value=item["Preco_Unit"])
            # Total como FORMULA (Quantidade * Preço Unitário), nao valor fixo --
            # assim a planilha recalcula sozinha se o usuario editar algo depois.
            ws.cell(row=linha_atual, column=4, value=f"=B{linha_atual}*C{linha_atual}")

            for col_idx in range(1, 5):
                celula = ws.cell(row=linha_atual, column=col_idx)
                celula.font = Font(name=fonte_padrao, size=10)
                celula.fill = PatternFill("solid", fgColor=zebra)
                celula.border = borda_fina
                if col_idx in (3, 4):
                    celula.number_format = 'R$ #,##0.00'
                if col_idx == 2:
                    celula.alignment = Alignment(horizontal="center")

            linha_atual += 1

        ultima_linha_grupo = linha_atual - 1

        # --- Subtotal da secao ---
        ws.merge_cells(f"A{linha_atual}:C{linha_atual}")
        celula_label = ws.cell(row=linha_atual, column=1, value=f"Subtotal — {tipo}")
        celula_label.alignment = Alignment(horizontal="right", vertical="center")
        celula_label.font = Font(name=fonte_padrao, size=10, bold=True, italic=True)
        celula_subtotal = ws.cell(
            row=linha_atual, column=4,
            value=f"=SUM(D{primeira_linha_grupo}:D{ultima_linha_grupo})"
        )
        celula_subtotal.font = Font(name=fonte_padrao, size=10, bold=True)
        celula_subtotal.number_format = 'R$ #,##0.00'
        linhas_subtotal.append(linha_atual)
        linha_atual += 2  # deixa uma linha em branco entre secoes

    # --- Linha de CUSTO DIRETO (soma dos subtotais de cada secao) ---
    linha_custo_direto = linha_atual
    label_custo = "CUSTO DIRETO" if bdi_percentual else "TOTAL ESTIMADO"
    ws.merge_cells(f"A{linha_custo_direto}:C{linha_custo_direto}")
    celula_label = ws.cell(row=linha_custo_direto, column=1, value=label_custo)
    celula_label.alignment = Alignment(horizontal="right", vertical="center")
    celula_label.fill = PatternFill("solid", fgColor=azul_cabecalho)
    celula_label.font = Font(name=fonte_padrao, size=12, bold=True, color="FFFFFF")

    formula_custo_direto = "+".join(f"D{linha}" for linha in linhas_subtotal)
    celula_custo_direto = ws.cell(row=linha_custo_direto, column=4, value=f"={formula_custo_direto}")
    celula_custo_direto.font = Font(name=fonte_padrao, size=12, bold=True, color="FFFFFF")
    celula_custo_direto.fill = PatternFill("solid", fgColor=azul_cabecalho)
    celula_custo_direto.number_format = 'R$ #,##0.00'
    ws.row_dimensions[linha_custo_direto].height = 22

    # --- Bloco de BDI (so aparece se um percentual > 0 foi informado) ---
    if bdi_percentual:
        linha_bdi = linha_custo_direto + 1
        ws.merge_cells(f"A{linha_bdi}:C{linha_bdi}")
        celula_bdi_label = ws.cell(
            row=linha_bdi, column=1,
            value=f"BDI ({bdi_percentual:g}%) — administração, lucro, impostos e imprevistos"
        )
        celula_bdi_label.alignment = Alignment(horizontal="right", vertical="center")
        celula_bdi_label.font = Font(name=fonte_padrao, size=10, italic=True)
        celula_bdi_valor = ws.cell(
            row=linha_bdi, column=4,
            value=f"=D{linha_custo_direto}*{bdi_percentual}/100"
        )
        celula_bdi_valor.font = Font(name=fonte_padrao, size=10)
        celula_bdi_valor.number_format = 'R$ #,##0.00'

        linha_preco_venda = linha_bdi + 1
        ws.merge_cells(f"A{linha_preco_venda}:C{linha_preco_venda}")
        celula_venda_label = ws.cell(row=linha_preco_venda, column=1, value="PREÇO DE VENDA")
        celula_venda_label.alignment = Alignment(horizontal="right", vertical="center")
        celula_venda_label.fill = PatternFill("solid", fgColor=verde_venda)
        celula_venda_label.font = Font(name=fonte_padrao, size=12, bold=True, color="FFFFFF")

        celula_venda_valor = ws.cell(
            row=linha_preco_venda, column=4,
            value=f"=D{linha_custo_direto}+D{linha_bdi}"
        )
        celula_venda_valor.font = Font(name=fonte_padrao, size=12, bold=True, color="FFFFFF")
        celula_venda_valor.fill = PatternFill("solid", fgColor=verde_venda)
        celula_venda_valor.number_format = 'R$ #,##0.00'
        ws.row_dimensions[linha_preco_venda].height = 22

    # --- Largura das colunas ---
    larguras = {1: 34, 2: 14, 3: 20, 4: 18}
    for col_idx, largura in larguras.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    wb.save(xlsx_path)
    return xlsx_path
